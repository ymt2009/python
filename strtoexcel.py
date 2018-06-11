#!/usr/bin/python
#coding:utf-8
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
#from openpyxl.writer.excel import ExcelWriter

import json
import sys,os
import re
reload(sys)#need reload sys
sys.setdefaultencoding('utf-8')

wb= Workbook()


def add_key(obj,ids,str):
    if('Null' == str):
        return
    else:
        obj[ids]= str


def writeFile():
    path_filename= r'result/strings.xlsx'
    ws= wb.worksheets[0]
    ws.title= 'strings'
    n= 1
    for k,v in content.iteritems():
        for x in range(index, 3):
            if x == 1 and index == 1:
                ws.cell(row=n,column= x).value= k
            if x == 2 and index == 1:
                ws.cell(row=n,column= x).value= v
        n+= 1

    m= 1
    for k,v in content.iteritems():
        for x in range(3, len(sys.argv)+2):
            if index == x-1:
                ws.cell(row=m,column= x).value= v
        m+= 1
    wb.save(filename= path_filename)

get_ids= {}
def file_parse(singleLanguage):
    global content, get_ids
    if not os.path.exists(singleLanguage+'.json'):
        return False
    f = open(singleLanguage+'.json', 'r')
    content= json.load(f)
    n= 0
    for k,v in content.iteritems():
        n+= 1
        add_key(get_ids, k, n)
    f.close()
    writeFile()


def eachLang(input_list):
    global index
    index= 0
    for lang in input_list:
        index+= 1
        file_parse(lang)
    wb.close()

def main(input_list):
    print input_list
    delpy= input_list.pop(0)
    eachLang(input_list)

if __name__ == '__main__':
    print '---Program start'
    if (len(sys.argv) < 2):
        print '-------Error: need write language name eg:en fr sp...'
    else:
        main(sys.argv)
